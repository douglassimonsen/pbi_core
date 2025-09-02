import json
from dataclasses import dataclass
from typing import TYPE_CHECKING

import openpyxl

from pbi_core.logging import get_logger
from pbi_core.static_files import Layout
from pbi_core.static_files.layout.sources.literal import LiteralSource, serialize_literal

if TYPE_CHECKING:
    from _typeshed import StrPath
logger = get_logger()


@dataclass
class TextElement:
    category: str
    xpath: list[str | int]
    field: str
    text: str


class TextElements:
    language: str = "source_text"
    text_elements: list[TextElement]

    def __init__(self, text_elements: list[TextElement] | None = None, language: str = "source_text") -> None:
        self.text_elements = text_elements or []
        self.language = language

    def to_csv(self) -> None:
        pass

    def _grouped(self) -> dict[str, list[TextElement]]:
        grouped: dict[str, list[TextElement]] = {}
        for element in self.text_elements:
            grouped.setdefault(element.category, []).append(element)
        return grouped

    def set_elements(self, layout: Layout) -> None:
        """Updates the text elements."""
        for text_element in self.text_elements:
            node = layout.find_xpath(text_element.xpath)
            if isinstance(node, LiteralSource):
                node.Literal.Value = serialize_literal(text_element.text)
            else:
                setattr(node, text_element.field, text_element.text)

    def to_excel(self, path: "StrPath") -> None:
        wb = openpyxl.Workbook()
        for category, objects in self._grouped().items():
            ws = wb.create_sheet(category)
            for j, name in enumerate(["xpath", "field", "source_text"]):
                ws.cell(1, j + 1).value = name
            for i, obj in enumerate(objects):
                ws.cell(2 + i, 1).value = json.dumps(obj.xpath)
                ws.cell(2 + i, 2).value = obj.field
                ws.cell(2 + i, 3).value = obj.text
        wb.remove(wb["Sheet"])
        wb.save(path)

    @classmethod
    def from_excel(cls, path: "StrPath") -> "list[TextElements]":
        logger.info("Parsing Excel", path=path)
        wb = openpyxl.load_workbook(path)

        headers = next(wb.worksheets[0].rows)
        languages: list[str] = [str(c.value) for c in headers][3:]

        static_elements: dict[str, list[TextElement]] = {}
        for ws in wb.worksheets:
            for row in list(ws.values)[1:]:
                assert isinstance(row[0], str)
                assert isinstance(row[1], str)
                assert isinstance(row[2], str)
                xpath = json.loads(row[0])
                field = row[1]
                for i, lang in enumerate(languages):
                    text = row[3 + i]
                    assert isinstance(text, str)
                    static_elements.setdefault(lang, []).append(
                        TextElement(
                            category=ws.title,
                            xpath=xpath,
                            field=field,
                            text=text,
                        ),
                    )
        return [
            TextElements(
                text_elements=elements,
                language=lang,
            )
            for lang, elements in static_elements.items()
        ]
