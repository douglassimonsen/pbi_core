import json
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from pbi_core import LocalReport
from pbi_core.logging import get_logger
from pbi_core.pydantic.main import BaseValidation
from pbi_core.static_files import Layout
from pbi_core.static_files.layout.filters import Filter
from pbi_core.static_files.layout.sources.literal import LiteralSource, serialize_literal
from pbi_core.static_files.layout.sources.paragraphs import TextRun
from pbi_core.static_files.layout.visuals.base import PropertyDef
from pbi_core.static_files.layout.visuals.properties.base import LiteralExpression
from pbi_core.static_files.layout.visuals.text_box import TextBox

if TYPE_CHECKING:
    from _typeshed import StrPath

logger = get_logger()


@dataclass
class StaticElement:
    category: str
    xpath: list[str | int]
    field: str
    text: str


class StaticElements:
    language: str = "source_text"
    static_elements: list[StaticElement]

    def __init__(self, static_elements: list[StaticElement] | None = None, language: str = "source_text") -> None:
        self.static_elements = static_elements or []
        self.language = language

    def to_csv(self) -> None:
        pass

    def _grouped(self) -> dict[str, list[StaticElement]]:
        grouped: dict[str, list[StaticElement]] = {}
        for element in self.static_elements:
            grouped.setdefault(element.category, []).append(element)
        return grouped

    def set_elements(self, layout: Layout) -> None:
        """Updates the static elements."""
        for static_element in self.static_elements:
            node = layout.find_xpath(static_element.xpath)
            if isinstance(node, LiteralSource):
                node.Literal.Value = serialize_literal(static_element.text)
            else:
                setattr(node, static_element.field, static_element.text)

    def to_excel(self, path: "StrPath") -> None:
        wb = openpyxl.Workbook()
        for category, objects in self._grouped().items():
            ws = wb.create_sheet(category)
            for j, name in enumerate(["xpath", "field", "source_text"]):
                ws.cell(1, j + 1).value = name
            for i, obj in enumerate(objects):
                ws.cell(2 + i, 1).value = json.dumps(obj.xpath)
                ws.cell(2 + i, 2).value = obj.field
                ws.cell(2 + i, 3).value = obj.text
        wb.remove(wb["Sheet"])
        wb.save(path)

    @classmethod
    def from_excel(cls, path: "StrPath") -> "list[StaticElements]":
        logger.info("Parsing Excel", path=path)
        wb = openpyxl.load_workbook(path)

        headers = next(wb.worksheets[0].rows)
        languages: list[str] = [str(c.value) for c in headers][3:]

        static_elements: dict[str, list[StaticElement]] = {}
        for ws in wb.worksheets:
            for row in list(ws.values)[1:]:
                assert isinstance(row[0], str)
                assert isinstance(row[1], str)
                assert isinstance(row[2], str)
                xpath = json.loads(row[0])
                field = row[1]
                for i, lang in enumerate(languages):
                    text = row[3 + i]
                    assert isinstance(text, str)
                    static_elements.setdefault(lang, []).append(
                        StaticElement(
                            category=ws.title,
                            xpath=xpath,
                            field=field,
                            text=text,
                        ),
                    )
        return [
            StaticElements(
                static_elements=elements,
                language=lang,
            )
            for lang, elements in static_elements.items()
        ]


def parse_config(config: BaseValidation | dict | None) -> list[LiteralSource]:
    if config is None:
        return []
    ret: list[LiteralSource] = []
    if isinstance(config, dict):
        for value in config.values():
            for element in value:
                assert isinstance(element, PropertyDef)
                assert isinstance(element.properties, dict)
                ret.extend(
                    field_value.expr
                    for field_value in element.properties.values()
                    if isinstance(field_value, LiteralExpression) and isinstance(field_value.expr.value(), str)
                )
    else:
        for field_name in config.__pydantic_fields__:
            value: list[BaseValidation] = getattr(config, field_name)
            for element in value:
                properties: BaseValidation = element.properties  # type: ignore reportAttributeAccessIssue
                for prop_name in properties.__pydantic_fields__:
                    prop_value = getattr(properties, prop_name)
                    if isinstance(prop_value, LiteralExpression) and isinstance(prop_value.expr.value(), str):
                        ret.append(prop_value.expr)
    return ret


def get_text_runs(viz: TextBox) -> list[TextRun]:
    ret = []
    for property_list in viz.objects.general:
        for paragraph in property_list.properties.paragraphs or []:
            ret.extend(run for run in paragraph.textRuns)
    return ret


def get_static_elements(layout: Layout) -> StaticElements:
    """Retrieves all static elements of a report Layout.

    The static elements in the report are:
    1. Section names
    2. Filter names
    3. Visual config names such as title, header, etc.
    """
    elements = []

    elements.extend(
        StaticElement(
            category="Filter",
            xpath=f.get_xpath(layout),
            field="displayName",
            text=f.displayName,
        )
        for f in layout.find_all(Filter)
        if f.displayName is not None
    )

    for section in layout.sections:
        elements.append(
            StaticElement(
                category="Section",
                xpath=section.get_xpath(layout),
                field="displayName",
                text=section.displayName,
            ),
        )

        for visual_container in section.visualContainers:
            for visual in visual_container.get_visuals():
                if isinstance(visual, TextBox):
                    text_runs = get_text_runs(visual)
                    elements.extend(
                        StaticElement(
                            category="Visual",
                            xpath=run.get_xpath(layout),
                            field="value",
                            text=run.value,
                        )
                        for run in text_runs
                        if isinstance(run.value, str)
                    )
                for config_obj in [visual.vcObjects, visual.objects]:
                    text_config = parse_config(config_obj)
                    for prop in text_config:
                        val = prop.value()
                        assert isinstance(
                            val,
                            str,
                        )  # technically done before, but the type checker doesn't remember that
                        elements.append(
                            StaticElement(
                                category="Visual",
                                xpath=prop.get_xpath(layout),
                                field="expr",
                                text=val,
                            ),
                        )

    return StaticElements(elements)


def set_static_elements(translation_path: "StrPath", pbix_path: "StrPath") -> None:
    """We parse an excel file containing translations and create a pbix file for each language.

    The excel file must have the following structure:
    - Each worksheet represents a category (e.g., "Section", "Visual", etc.)
    - The first row contains the headers: "xpath", "field", "default", and then one column for each language
        (e.g., "en", "fr", "de").
    """
    wb = openpyxl.load_workbook(translation_path)
    languages: list[str] = [str(x) for x in next(iter(wb.worksheets[0].values))[3:]]
    processing: dict[str, list[StaticElement]] = {}
    for ws in wb.worksheets:
        for row in list(ws.values)[1:]:
            for i, language in enumerate(languages):
                processing.setdefault(language, []).append(
                    StaticElement(
                        category=ws.title,
                        xpath=json.loads(str(row[0])),
                        field=str(row[1]),
                        text=str(row[3 + i]),
                    ),
                )
    for language, static_elements in processing.items():
        pbix = LocalReport.load_pbix(pbix_path)
        for static_element in static_elements:
            node = pbix.static_files.layout.find_xpath(static_element.xpath)
            setattr(node, static_element.field, static_element.text)
        out_path = f"{Path(pbix_path).with_suffix('').absolute().as_posix()}_{language}.pbix"
        pbix.save_pbix(out_path)
